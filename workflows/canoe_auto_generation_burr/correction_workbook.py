"""Correction workbook / markdown generator.

Translates correction items produced by the test-case validation step into
human-friendly outputs:

* ``generate_correction_workbook`` – writes an ``.xlsx`` file so users can see
  exactly which row/column needs fixing, with error rows highlighted in red
  and warning rows in yellow (when *openpyxl* is available; otherwise a
  minimal spreadsheet is written via ``zipfile`` + raw OpenXML).
* ``generate_correction_markdown`` – writes a markdown table grouped by
  severity (errors first, then warnings).

The module mirrors the zero-dependency OpenXML approach used by the existing
template parser in ``canoe_workflow.py``: reading and writing ``.xlsx`` files
is done with the standard library only.
"""

from __future__ import annotations

import html
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence
from xml.sax.saxutils import escape as _xml_escape

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

COLUMNS: List[str] = ["用例ID", "步骤序号", "严重级别", "类别", "问题描述", "建议修复"]

_OPENPYXL_AVAILABLE: Optional[bool] = None


def _openpyxl_is_available() -> bool:
    """Return ``True`` when *openpyxl* can be imported."""

    global _OPENPYXL_AVAILABLE
    if _OPENPYXL_AVAILABLE is None:
        try:
            import openpyxl  # noqa: F401

            _OPENPYXL_AVAILABLE = True
        except ImportError:
            _OPENPYXL_AVAILABLE = False
    return _OPENPYXL_AVAILABLE


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _as_text(value: Any) -> str:
    """Convert *value* to a clean string (mirrors ``canoe_workflow._as_text``)."""

    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _item_to_row(item: Dict[str, Any]) -> List[str]:
    """Map a correction item dict to the display column order."""

    severity = _as_text(item.get("severity"))
    suggestion = _as_text(item.get("suggestion")) or _as_text(item.get("fix"))
    if not suggestion:
        field = _as_text(item.get("field"))
        if field:
            suggestion = f"请检查「{field}」列的填写内容"
        else:
            suggestion = "请对照模板要求修正该行数据"
    return [
        _as_text(item.get("test_case_id")),
        _as_text(item.get("step_no")),
        severity or "warning",
        _as_text(item.get("category")),
        _as_text(item.get("message")),
        suggestion,
    ]


def _sort_items(items: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Sort items so that errors come before warnings."""

    def severity_rank(item: Dict[str, Any]) -> int:
        sev = _as_text(item.get("severity")).lower()
        if sev == "error":
            return 0
        if sev == "warning":
            return 1
        return 2

    return sorted(items, key=severity_rank)


def _ensure_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _approx_col_width(rows: Sequence[Sequence[str]], col_idx: int) -> int:
    """Approximate the max display width for column *col_idx*."""

    max_len = len(COLUMNS[col_idx]) if col_idx < len(COLUMNS) else 10
    for row in rows:
        cell = row[col_idx] if col_idx < len(row) else ""
        # CJK characters are roughly double-width
        visual_len = sum(2 if ord(ch) > 0x2E80 else 1 for ch in str(cell))
        if visual_len > max_len:
            max_len = visual_len
    return min(max_len + 4, 60)


# ---------------------------------------------------------------------------
# Minimal .xlsx writer (no openpyxl dependency)
# ---------------------------------------------------------------------------

# OpenXML namespace constants
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"

# Cell type code mapping for the ``t`` attribute
_TYPE_SHARED = "s"  # shared string


def _col_letter(col_idx: int) -> str:
    """Convert a 0-based column index to an Excel column letter (A, B, ...)."""

    result = ""
    n = col_idx + 1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _write_minimal_xlsx(
    output_path: Path,
    header_row: Sequence[str],
    data_rows: Sequence[Sequence[str]],
    *,
    error_row_indices: Optional[set[int]] = None,
    warning_row_indices: Optional[set[int]] = None,
) -> None:
    """Write a minimal ``.xlsx`` file using only ``zipfile`` and XML strings.

    Parameters
    ----------
    output_path
        Destination file path.
    header_row
        Column headers.
    data_rows
        Data rows (each a list of cell strings).
    error_row_indices
        0-based indices into *data_rows* that should be flagged as errors.
    warning_row_indices
        0-based indices into *data_rows* that should be flagged as warnings.
    """

    _ensure_dir(output_path)
    error_rows = error_row_indices or set()
    warning_rows = warning_row_indices or set()

    # Build shared strings table (header + all data cells)
    all_strings: List[str] = []
    index_map: Dict[str, int] = {}

    def _shared_index(text: str) -> int:
        if text not in index_map:
            index_map[text] = len(all_strings)
            all_strings.append(text)
        return index_map[text]

    # Build all rows
    num_cols = len(header_row)

    # Sheet XML
    sheet_parts: List[str] = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<worksheet xmlns="{_NS_MAIN}"'
        f' xmlns:r="{_NS_R}">',
    ]

    # Column widths
    all_cell_rows = [list(header_row)] + [list(r) for r in data_rows]
    for ci in range(num_cols):
        width = _approx_col_width(all_cell_rows, ci)
        # Convert approximate width to Excel column width units
        col_w = max(width * 0.9, 8)
        sheet_parts.append(
            f'<col min="{ci + 1}" max="{ci + 1}" width="{col_w:.2f}" customWidth="1"/>'
        )

    # Sheet data
    sheet_parts.append(f'<sheetData>')

    # Header row (row 1)
    sheet_parts.append('<row r="1">')
    for ci, val in enumerate(header_row):
        ref = f"{_col_letter(ci)}1"
        si = _shared_index(str(val))
        sheet_parts.append(
            f'<c r="{ref}" s="1" t="{_TYPE_SHARED}"><v>{si}</v></c>'
        )
    sheet_parts.append('</row>')

    # Data rows
    for ri, row_data in enumerate(data_rows):
        excel_row = ri + 2  # 1-based, header is row 1
        style_id = 2 if ri in error_rows else (3 if ri in warning_rows else 0)
        sheet_parts.append(f'<row r="{excel_row}">')
        for ci in range(num_cols):
            val = str(row_data[ci]) if ci < len(row_data) else ""
            ref = f"{_col_letter(ci)}{excel_row}"
            si = _shared_index(val)
            cell_parts = [f'<c r="{ref}"']
            if style_id:
                cell_parts.append(f' s="{style_id}"')
            cell_parts.append(f' t="{_TYPE_SHARED}"><v>{si}</v></c>')
            sheet_parts.append("".join(cell_parts))
        sheet_parts.append('</row>')

    sheet_parts.append('</sheetData>')
    sheet_parts.append('</worksheet>')

    sheet_xml = "".join(sheet_parts)

    # Shared strings XML
    ss_parts: List[str] = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<sst xmlns="{_NS_MAIN}"'
        f' count="{sum(1 for r in all_cell_rows for _ in r)}"'
        f' uniqueCount="{len(all_strings)}">',
    ]
    for s in all_strings:
        ss_parts.append(f'<si><t xml:space="preserve">{_xml_escape(s)}</t></si>')
    ss_parts.append("</sst>")
    shared_strings_xml = "".join(ss_parts)

    # Styles XML (style 0 = default, 1 = header bold, 2 = error fill, 3 = warning fill)
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<styleSheet xmlns="{_NS_MAIN}">'
        '<fonts count="2">'
        '<font><sz val="11"/><name val="Calibri"/></font>'
        '<font><b/><sz val="11"/><name val="Calibri"/></font>'
        '</fonts>'
        '<fills count="4">'
        '<fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="gray125"/></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FFFFC7CE"/></patternFill></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FFFFEB9C"/></patternFill></fill>'
        '</fills>'
        '<borders count="1"><border/></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="4">'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
        '<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1"/>'
        '<xf numFmtId="0" fontId="0" fillId="2" borderId="0" xfId="0" applyFill="1"/>'
        '<xf numFmtId="0" fontId="0" fillId="3" borderId="0" xfId="0" applyFill="1"/>'
        '</cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        '</styleSheet>'
    )

    # Workbook XML
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<workbook xmlns="{_NS_MAIN}"'
        f' xmlns:r="{_NS_R}">'
        '<sheets><sheet name="Corrections" sheetId="1" r:id="rId1"/></sheets>'
        '</workbook>'
    )

    # Workbook rels XML
    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{_NS_PKG_REL}">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '</Relationships>'
    )

    # Workbook internal rels
    sheet_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{_NS_PKG_REL}">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '</Relationships>'
    )

    # Content types XML
    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '</Types>'
    )

    # Write the zip archive
    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml)
        zf.writestr("_rels/.rels", workbook_rels_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", sheet_rels_xml)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
        zf.writestr("xl/sharedStrings.xml", shared_strings_xml)
        zf.writestr("xl/styles.xml", styles_xml)


# ---------------------------------------------------------------------------
# openpyxl-based writer (richer formatting)
# ---------------------------------------------------------------------------

def _write_openpyxl_xlsx(
    output_path: Path,
    header_row: Sequence[str],
    data_rows: Sequence[Sequence[str]],
    *,
    error_row_indices: Optional[set[int]] = None,
    warning_row_indices: Optional[set[int]] = None,
) -> None:
    """Write an ``.xlsx`` file using *openpyxl* with rich formatting."""

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _ensure_dir(output_path)
    error_rows = error_row_indices or set()
    warning_rows = warning_row_indices or set()

    wb = Workbook()
    ws = wb.active
    ws.title = "Corrections"

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    for ci, val in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=ci, value=str(val))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Error / warning fills
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # Data rows
    for ri, row_data in enumerate(data_rows):
        excel_row = ri + 2
        fill = None
        if ri in error_rows:
            fill = error_fill
        elif ri in warning_rows:
            fill = warning_fill
        for ci in range(len(header_row)):
            val = str(row_data[ci]) if ci < len(row_data) else ""
            cell = ws.cell(row=excel_row, column=ci + 1, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    # Auto-fit column widths (approximate)
    for ci in range(len(header_row)):
        max_len = len(str(header_row[ci]))
        for ri in range(len(data_rows)):
            val = str(data_rows[ri][ci]) if ci < len(data_rows[ri]) else ""
            visual_len = sum(2 if ord(ch) > 0x2E80 else 1 for ch in val)
            if visual_len > max_len:
                max_len = visual_len
        col_letter = get_column_letter(ci + 1)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_correction_workbook(
    correction_items: List[Dict[str, Any]],
    structured_test_case: Dict[str, Any],
    output_path: Path,
) -> Dict[str, Any]:
    """Generate an Excel correction workbook from correction items.

    Each correction item becomes a row in the spreadsheet with columns:
    用例ID, 步骤序号, 严重级别, 类别, 问题描述, 建议修复.

    Error rows are highlighted in red and warning rows in yellow when
    *openpyxl* is available.  When *openpyxl* is not installed, a minimal
    ``.xlsx`` file is written using only the standard library.

    Parameters
    ----------
    correction_items
        List of correction item dicts produced by the test-case validation
        step.  Each item has keys: ``severity``, ``category``,
        ``test_case_id``, ``step_no``, ``field``, ``message``, and
        optionally ``suggestion`` or ``fix``.
    structured_test_case
        The structured test case dict (used for metadata; currently only
        the project name is extracted for the sheet title context).
    output_path
        Destination ``.xlsx`` file path.

    Returns
    -------
    dict
        ``{"path": str(output_path), "row_count": N, "status": "generated"}``
    """

    output_path = Path(output_path)

    # structured_test_case is accepted for API symmetry with the markdown
    # generator; the project name is available for future metadata use.
    if isinstance(structured_test_case, dict):
        _ = structured_test_case.get("project", {}).get("name", "")

    sorted_items = _sort_items(correction_items)
    data_rows = [_item_to_row(item) for item in sorted_items]

    error_indices: set[int] = set()
    warning_indices: set[int] = set()
    for idx, item in enumerate(sorted_items):
        sev = _as_text(item.get("severity")).lower()
        if sev == "error":
            error_indices.add(idx)
        elif sev == "warning":
            warning_indices.add(idx)

    if _openpyxl_is_available():
        _write_openpyxl_xlsx(
            output_path,
            COLUMNS,
            data_rows,
            error_row_indices=error_indices,
            warning_row_indices=warning_indices,
        )
    else:
        _write_minimal_xlsx(
            output_path,
            COLUMNS,
            data_rows,
            error_row_indices=error_indices,
            warning_row_indices=warning_indices,
        )

    return {
        "path": str(output_path),
        "row_count": len(correction_items),
        "status": "generated",
    }


def generate_correction_markdown(
    correction_items: List[Dict[str, Any]],
    structured_test_case: Dict[str, Any],
    output_path: Path,
) -> Dict[str, Any]:
    """Generate a markdown correction report grouped by severity.

    Parameters
    ----------
    correction_items
        List of correction item dicts.
    structured_test_case
        The structured test case dict (for metadata context).
    output_path
        Destination ``.md`` file path.

    Returns
    -------
    dict
        ``{"path": str(output_path), "row_count": N, "status": "generated"}``
    """

    output_path = Path(output_path)
    _ensure_dir(output_path)

    project_name = (
        structured_test_case.get("project", {}).get("name", "CANoe_AutoTest_Project")
        if isinstance(structured_test_case, dict)
        else "CANoe_AutoTest_Project"
    )

    sorted_items = _sort_items(correction_items)
    error_items = [it for it in sorted_items if _as_text(it.get("severity")).lower() == "error"]
    warning_items = [it for it in sorted_items if _as_text(it.get("severity")).lower() == "warning"]
    other_items = [
        it for it in sorted_items
        if _as_text(it.get("severity")).lower() not in ("error", "warning")
    ]

    lines: List[str] = []
    lines.append(f"# 测试用例修正报告 — {project_name}")
    lines.append("")
    lines.append(f"**修正项总数**: {len(correction_items)}  ")
    lines.append(f"**错误**: {len(error_items)} | **警告**: {len(warning_items)} | **其他**: {len(other_items)}")
    lines.append("")
    lines.append("---")
    lines.append("")

    def _write_section(title: str, icon: str, items: List[Dict[str, Any]]) -> None:
        if not items:
            return
        lines.append(f"## {icon} {title} ({len(items)})")
        lines.append("")
        lines.append("| 用例ID | 步骤序号 | 严重级别 | 类别 | 问题描述 | 建议修复 |")
        lines.append("|--------|----------|----------|------|----------|----------|")
        for item in items:
            row = _item_to_row(item)
            escaped = [html.escape(str(cell), quote=False) for cell in row]
            lines.append("| " + " | ".join(escaped) + " |")
        lines.append("")

    _write_section("错误 (Error)", "🔴", error_items)
    _write_section("警告 (Warning)", "🟡", warning_items)
    _write_section("其他 (Other)", "ℹ️", other_items)

    if not sorted_items:
        lines.append("## ✅ 无修正项")
        lines.append("")
        lines.append("所有测试用例均通过校验，无需修正。")

    output_path.write_text("\n".join(lines), encoding="utf-8")

    return {
        "path": str(output_path),
        "row_count": len(correction_items),
        "status": "generated",
    }
